import io
from pathlib import Path
from typing import Optional

import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.utils import get_column_letter
import httpx

from .scraper import (
    BASE_URL,
    TORGET_SUBCATEGORIES,
    detect_changes,
    extract_categories,
    extract_subcategories,
    fetch_html,
    scrape_listings,
)

app = FastAPI(title="Finn Parser")

BASE_DIR = Path(__file__).resolve().parent
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


async def _download_image(url: str, timeout_ms: int = 10000) -> Optional[bytes]:
    if not url:
        return None
    try:
        timeout = httpx.Timeout(timeout_ms / 1000)
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        }
        async with httpx.AsyncClient(timeout=timeout, headers=headers, follow_redirects=True) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            return resp.content
    except Exception:
        return None


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/categories")
async def get_categories(base_url: Optional[str] = None):
    url = base_url or BASE_URL
    html = await fetch_html(url)
    items = extract_categories(html, url)
    return {"items": [item.__dict__ for item in items]}


@app.get("/api/subcategories")
async def get_subcategories(category_url: str):
    html = await fetch_html(category_url)
    items = extract_subcategories(html, category_url)
    return {"items": [item.__dict__ for item in items]}


@app.get("/api/torget-subcategories")
async def get_torget_subcategories():
    return {"items": [{"name": name, "url": url} for name, url in TORGET_SUBCATEGORIES]}


@app.post("/api/parse")
async def parse_category(
    category_name: str = Form(""),
    subcategory_name: str = Form(""),
    subcategory_url: str = Form(""),
    max_items: int = Form(50),
    preview: int = Form(0),
):
    if not subcategory_url:
        raise HTTPException(status_code=400, detail="subcategory_url is required")

    listings = await scrape_listings(
        category_name=category_name or "",
        subcategory_name=subcategory_name or "",
        subcategory_url=subcategory_url,
        max_items=max_items,
    )

    rows = [listing.__dict__ for listing in listings]
    if not rows:
        return JSONResponse({"items": []})

    if preview:
        return JSONResponse({"items": rows})

    image_urls = [row.get("image", "") for row in rows]
    image_blobs = []
    for url in image_urls:
        image_blobs.append(await _download_image(url))

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="data")
        ws = writer.sheets["data"]
        headers = [cell.value for cell in ws[1]]
        if "image" in headers:
            image_col = headers.index("image") + 1
            col_letter = get_column_letter(image_col)
            ws.column_dimensions[col_letter].width = 15
            for idx, blob in enumerate(image_blobs, start=2):
                if not blob:
                    continue
                try:
                    img = XlsxImage(io.BytesIO(blob))
                    img.width = 80
                    img.height = 80
                    cell = f"{col_letter}{idx}"
                    ws.add_image(img, cell)
                    ws.row_dimensions[idx].height = 60
                except Exception:
                    continue
        meta_df = pd.DataFrame(
            [{
                "category_name": category_name,
                "subcategory_name": subcategory_name,
                "subcategory_url": subcategory_url,
            }]
        )
        meta_df.to_excel(writer, index=False, sheet_name="meta")

    output.seek(0)
    headers = {
        "Content-Disposition": "attachment; filename=finn_listings.xlsx"
    }
    return StreamingResponse(output, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/recheck")
async def recheck_table(
    file: UploadFile = File(...),
    max_items: int = Form(50),
):
    data = await file.read()
    try:
        xls = pd.ExcelFile(io.BytesIO(data))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid XLSX file") from exc

    if "data" not in xls.sheet_names:
        raise HTTPException(status_code=400, detail="XLSX must have a 'data' sheet")

    old_df = pd.read_excel(xls, sheet_name="data").fillna("")
    if old_df.empty:
        return JSONResponse({"message": "changes not found", "items": []})

    subcategory_url = ""
    if "meta" in xls.sheet_names:
        meta_df = pd.read_excel(xls, sheet_name="meta")
        if not meta_df.empty:
            subcategory_url = str(meta_df.iloc[0].get("subcategory_url", ""))

    if not subcategory_url:
        raise HTTPException(status_code=400, detail="subcategory_url not found in meta sheet")

    category_name = ""
    subcategory_name = ""
    if "meta" in xls.sheet_names:
        meta_df = pd.read_excel(xls, sheet_name="meta")
        if not meta_df.empty:
            category_name = str(meta_df.iloc[0].get("category_name", ""))
            subcategory_name = str(meta_df.iloc[0].get("subcategory_name", ""))

    new_listings = await scrape_listings(
        category_name=category_name,
        subcategory_name=subcategory_name,
        subcategory_url=subcategory_url,
        max_items=max_items,
    )
    new_rows = [listing.__dict__ for listing in new_listings]

    changed_rows = detect_changes(old_df.to_dict(orient="records"), new_rows)
    if not changed_rows:
        return JSONResponse({"message": "changes not found", "items": []})

    changed_df = pd.DataFrame(changed_rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        changed_df.to_excel(writer, index=False, sheet_name="changes")

    output.seek(0)
    headers = {
        "Content-Disposition": "attachment; filename=finn_changes.xlsx"
    }
    return StreamingResponse(output, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
